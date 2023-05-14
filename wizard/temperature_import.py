# models/temperature_log_import_wizard.py
import base64
from openpyxl import load_workbook
from io import BytesIO
from odoo import models, fields, exceptions


class TemperatureLogImportWizard(models.TransientModel):
    _name = 'temperature.log.import.wizard'
    _description = 'Temperature Log Import Wizard'

    temperature_logger_id = fields.Many2one('temperature.logger', string='Temperature Logger', required=True)
    data_file = fields.Binary('Excel File', required=True)
    filename = fields.Char('Filename')

    def import_temperature_log(self):
        # Check if the uploaded file is present
        if not self.data_file:
            raise exceptions.UserError('No file found for import.')

        # Read the uploaded file and load the workbook
        excel_file_data = base64.b64decode(self.data_file)
        workbook = load_workbook(BytesIO(excel_file_data))

        # Select the first sheet in the workbook (assuming data is in the first sheet)
        sheet = workbook.active

        # Iterate through the rows in the sheet, skipping the header row
        for row in sheet.iter_rows(min_row=2):
            time, temperature, humidity = row

            # Create a new temperature.log record with the parsed data
            self.env['temperature.log'].create({
                'temperature_logger_id': self.temperature_logger_id.id,
                'time': time.value,
                'temperature': temperature.value,
                'humidity': humidity.value
            })

        # Close the workbook
        workbook.close()